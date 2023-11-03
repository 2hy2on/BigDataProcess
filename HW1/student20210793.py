#!/usr/bin/python3

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("student.xlsx")
sheet = wb['Sheet1']
result_list = []

for i in range(2,sheet.max_row + 1):
    total = 0
    row_list=[]
    for j in range(1, 7):
        val = sheet.cell(row=i, column=j).value
        if j == 3 : ##midterm임
            total += float(val) * 0.3
        elif j == 4: ##final 이면
            total += float(val) * 0.35
        elif j == 5: ##homework이면
            total += float(val) * 0.34
        elif j == 6: ##attendence
            total += float(val)
        row_list.append(val)
        
    # print(sum) ##이거 h에 넣어야함
    sheet.cell(row=i, column=7, value=total) #total
    row_list.append(total)
    result_list.append(row_list)

result_list.sort(key=lambda x: x[6], reverse=True)

##만점자 처리
def check_perfect_score():
    perfect_num = 0
    for n in range(len(result_list)):
        if result_list[n][6] == 100:
            perfect_num += 1
    return perfect_num

def check_same_score(old_boundary, new_boundary):
    targetValue = result_list[old_boundary][6]
    if targetValue == 100:
        return (old_boundary, new_boundary)
    targetCount = -1
    for i, row in enumerate(result_list):
        if row[6] == targetValue:
            targetCount += 1
    print(targetCount)
    print("old_boundary__", old_boundary)
    print("new_boundary__", new_boundary)
    if old_boundary < targetCount or new_boundary < targetCount:
        return (old_boundary, new_boundary)

    old_boundary -= targetCount
    new_boundary -= targetCount    
    return (old_boundary, new_boundary)
    
perfect_cnt = check_perfect_score()
a_cutline = int(sheet.max_row * 0.3)
aa_cutline = int(a_cutline * 0.5)
b_cutline = int(sheet.max_row * 0.7)
bb_cutline = a_cutline+ int((b_cutline- a_cutline) * 0.5)

##만점자 처리
if perfect_cnt > aa_cutline and perfect_cnt <= a_cutline:
    aa_cutline = 0
elif perfect_cnt > a_cutline and perfect_cnt <= bb_cutline:
    aa_cutline = 0
    a_cutline = 0
elif perfect_cnt > bb_cutline and perfect_cnt <= b_cutline:
    aa_cutline = 0
    a_cutline = 0
    bb_cutline = 0

aa_cutline, a_cutline = check_same_score(aa_cutline, a_cutline)
a_cutline, bb_cutline = check_same_score(a_cutline, bb_cutline)
bb_cutline, b_cutline = check_same_score(bb_cutline, b_cutline)

#F주기
for i in range(len(result_list)):
    if result_list[i][6] < 40:
       result_list[i].append('F')

##A+주기
for i in range(aa_cutline):
    if result_list[i][6] >= 40:
        result_list[i].append('A+')
##A주기
for i in range(aa_cutline, a_cutline):
    if result_list[i][6] >= 40:
        result_list[i].append('A')
##B+주기
for i in range(a_cutline, bb_cutline):
    if result_list[i][6] >= 40:
        result_list[i].append('B+')
##B주기
for i in range(bb_cutline, b_cutline):
    if result_list[i][6] >= 40:
        result_list[i].append('B')
#print(result_list[b_cutline][6])

c_count = 0

for i in range(b_cutline, len(result_list)):
    if result_list[i][6] >= 40:
        c_count += 1
        result_list[i].append('C')

cc_count = c_count // 2
# print(b_cutline+cc_count+1)
if perfect_cnt <= b_cutline+cc_count:
    for i in range(b_cutline, b_cutline+cc_count):
        result_list[i].remove('C')
        result_list[i].append('C+')


result_list.sort(key=lambda x: x[0]) ##원상복귀

for i in range(2, sheet.max_row+1):
     for j in range(1,9):
        sheet.cell(row=i, column=j, value=result_list[i-2][j-1])

wb.save("student.xlsx")